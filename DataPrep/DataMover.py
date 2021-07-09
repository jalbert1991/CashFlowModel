# -*- coding: utf-8 -*-
"""
Created on Thu Feb 13 15:14:12 2020

@author: jalbert
"""

import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)

import sqlalchemy as sa
from sqlalchemy.orm import sessionmaker
import pandas as pd
import numpy as np
import warnings
import itertools
import openpyxl as xl
import os
import pathlib
import io
import traceback
import re
import copy
import time
import datetime as date
import json
import logging

from dateutil import relativedelta

#import project modules
#import GUI.DataPrepGUI as gui

class SQLEngine(object):
    """
    Class to create and manage sql server connections
    """
    def __init__(self, server_name='srv-produs-dw', db_name='edw_prod'):
        full_string = 'mssql+pyodbc://{}/{}?driver=ODBC+Driver+13+for+SQL+Server'.format(server_name, db_name)
        self.engine = sa.create_engine(full_string, fast_executemany=True) #echo=True
    def open_conn(self):
        self.conn = self.engine.raw_connection()
        self.cursor = self.conn.cursor()
    def close_conn(self):
        self.conn.close()
        self.engine.dispose()   
        
    def execute(self, sql_cmd, parameters=[], output=False):
        
        sql_cmd = 'set nocount on; ' + sql_cmd
        
        result_list=[]
        
        try:
            self.open_conn()

            self.cursor.execute(sql_cmd, *parameters)
            
            if output:
                rows = self.cursor.fetchall()
                cols = [row[0] for row in self.cursor.description]

                result_list.append(pd.DataFrame.from_records(rows, columns=cols)) #
                
                #check for any additional result sets
                while True:
                    try:
                        self.cursor.nextset()
                        
                        rows = self.cursor.fetchall()
                        cols = [row[0] for row in self.cursor.description]
                        result_list.append(pd.DataFrame.from_records(rows, columns=cols))
                    except:
                        #traceback.print_exc()
                        break
            
            self.cursor.commit()
            
        except Exception:
            traceback.print_exc()
        finally:
            self.close_conn()
            
        if output:
            #unpacking a tuple of 1 isnt super straight forward. so in this case just return the value as is
            if len(result_list)==1:
                return result_list[0]
            else:
                return tuple(result_list)
            return result_list 

      
    def upload_df(self, data_frame, schema, table_name, include_index=False, index_name=None, if_exists='fail'):
        """
        uploads a dataframe into a specified table in sql server
        """
        self.open_conn()
        try:
            data_frame.to_sql(name=table_name, con=self.engine, schema=schema, if_exists=if_exists, index=include_index, index_label=index_name)
        except Exception:
            traceback.print_exc()
        finally:
            self.close_conn()
                
class ExcelEngine(object):
    """
    class to manage excel workbook data transfers
    (only in a class to group functionality)
    """

    def load_workbook(self, raw_path):
        with open(raw_path, "rb") as f:
            in_mem_file = io.BytesIO(f.read())
        
        wb = xl.load_workbook(in_mem_file, data_only=True, read_only=True)
        return wb
    
    def load_worksheet_range(self, ws, bgn_rng='', end_rng=''):
        
        #parse begin end end range
        bgn_rng = re.findall(r'([a-zA-Z]*)(\d*)',bgn_rng)[0]
        end_rng = re.findall(r'([a-zA-Z]*)(\d*)',end_rng)[0]
        
        bgn_col = self.empty_to_none(bgn_rng[0])
        bgn_row = self.empty_to_none(bgn_rng[1])
        end_col = self.empty_to_none(end_rng[0])
        end_row = self.empty_to_none(end_rng[1])
        
        #convert inputs to int
        if isinstance(bgn_col,str):
            bgn_col = xl.utils.cell.column_index_from_string(bgn_col)
        if isinstance(end_col,str):
            end_col = xl.utils.cell.column_index_from_string(end_col)
        if bgn_row:
            bgn_row = int(bgn_row)
        if end_row:
            end_row = int(end_row)

        row_data = ws.iter_rows(min_col=bgn_col, min_row=bgn_row, max_col=end_col, max_row=end_row, values_only=True)
        row_values = [[cell for cell in row] for row in row_data]
        
        return pd.DataFrame(row_values[1:], columns=row_values[0])
        #return row_data
    
    def empty_to_none(self, text_sample):
        if str(text_sample).strip()=='':
            return None
        elif not text_sample:
            return None
        else:
            return str(text_sample)
    
    def unpivot_df(self, curve_df, pivot_col_name = 'vintage'):    
        total_col_list = curve_df.columns.values
        non_num_col = [x for x in total_col_list if not str(x).isdigit()]
        df_output = curve_df.melt(id_vars = non_num_col, var_name = pivot_col_name, value_name ='rate')
        return df_output
    
    def from_excel(self, file_path, ws_name, ws_range=None):
        """
        Load data from source excel file
        
        =======================
        Parameters:
        file_path: raw string
            full file path. enter with "r' prefix"
        ws_name: string
            worksheet name
        ws_range: string
            cell range containing data
        key_col: string
            column containing curve unique identifier 
            * this is renamed to "curve name" once loaded
        """
        wb = self.load_workbook(file_path)
        try:
            str_split = ws_range.split(':')
        except:
            str_split = ['','']
        cell_range = self.load_worksheet_range(wb[ws_name], str_split[0], str_split[1])
        return cell_range
    
        
class DataPrep(object):
    """
    Data Prep class imports model data from various sources
    
    Data Tape is the loan level monthly snapshot showing performance. 
        This data is where the model will begin calculations
    Data Curves are the transition probability curves to be utilized in the model
        CDR, CPR, Recovery, Roll Rates, etc. 
    """
    def __init__(self):
        self.sql_engine_import = SQLEngine(server_name='srv-produs-dw', db_name='edw_prod')
        self.sql_engine_export = SQLEngine(server_name='srv-produs-dw', db_name='FRA')
        self.excel_engine = ExcelEngine()
        #self.data_prep_gui = gui.DataPrepMain(self)
        
        #self.cutoff_date = None
        self.data_tape = None 
        self.prior_uw_projections = None
        
        self.rate_curve_groups={}
        self.model_configs={}
        self.index_projections={}
        
        self.segment_types=['default','prepay', 'curtail', 'recovery','rollrate', 'index']
        
    def import_index_projections(self, cutoff_date, projection_date=None):
        """
        Imports index projections from sql server tables

        Parameters
        ----------
        cutoff_date : date
            cutoff date for model scenario.
        projection_date : date
            optional. if provided will select a projection curve from a specific point in time
            i.e. backtest an older projection or forward fill with actuals up to a future projection date

        """
        sql_cmd = "exec fra.cf_model.usp_extract_index_proj ?, ?"
        params = [cutoff_date, projection_date]
        projections = self.sql_engine_import.execute(sql_cmd, params, output=True)
        
        self.index_projections[(projection_date, cutoff_date)] = IndexProjection(projections, projection_date, cutoff_date)
        
    def import_data_tape_cdw(self, deal_ids=[], batch_keys=[], asset_class=None, model_name=None):
        """
        Extracts snapshot data from the CDW. This is the "Current" data to project forward plus any historical data available.
        
        Parameters
        =================
        batch_ids: list of int(s)
            ids for specific batches in edw_prod.cdw.batch
        """
        
        deal_lst = ",".join([str(x) for x in deal_ids])
        batch_lst = ",".join([str(x) for x in batch_keys])
        
        sql_cmd = "exec fra.cf_model.usp_extract_snapshot_v2 ?, ?, ?"
        params = [deal_lst, batch_lst, asset_class]
        
        sql_tape = self.sql_engine_import.execute(sql_cmd, params, output=True)
        #self.data_tape = DataTape(sql_tape) #, cutoff_date)
        return DataTape(sql_tape)

    def import_data_tape_query(self, query): #cutoff_date=None
        """
        Extracts snapshot data from the custom sql query. This is the "Current" data to project forward.
        
        Parameters
        =================
        query: str
            sql query to download snapshot data
        """
        sql_tape = self.sql_engine_import.execute(query, output=True)

        #self.data_tape = DataTape(sql_tape) #cutoff_date
        return DataTape(sql_tape)
            
    def import_data_tape_excel(self, file_path, ws_name, ws_range):
        """
        imports snapshot data from excel file. This is the current data to project forward
        =======================
        Parameters:
        file_path: raw string
            full file path. enter with "r' prefix"
        ws_name: string
            worksheet name
        ws_range: string
            cell range containing data
        """
        #wb = self.excel_engine.load_workbook(file_path)
        #cell_range = self.excel_engine.load_worksheet_range(wb[ws_name], ws_range)
        #return cell_range
        
        pass
    
    def save_data_tape_source(self, model_name, dt_source):
        sql_cmd = """UPDATE fra.cf_model.model
                    SET data_tape_source = '{}'
                    WHERE 1=1
                    	AND model_name='{}'
                    """.format(dt_source, model_name)
                    
        self.sql_engine_export.execute(sql_cmd, output=False)
        
    
    def import_projections_cdw(self, deal_ids=[], batch_keys=[], projection_level='deal'):
        """
        Extracts prior projections data from the CDW. This is the "Current" data to project forward plus any historical data available.
        
        Parameters
        =================
        batch_ids: list of int(s)
            ids for specific batches in edw_prod.cdw.batch
        """
        
        deal_lst = ",".join([str(x) for x in deal_ids])
        batch_lst = ",".join([str(x) for x in batch_keys])
        
        sql_cmd = "exec fra.cf_model.usp_extract_projection ?, ?, ?"
        params = [deal_lst, batch_lst, projection_level]
        
        #self.prior_uw_projections = self.sql_engine_import.execute(sql_cmd, params, output=True)
        return self.sql_engine_import.execute(sql_cmd, params, output=True)
        
    def import_rate_curves_sql(self, curve_group, source_curve_group_name=None, model_name=None, scenario_name=None, curve_type='all', curve_sub_type='all', update_curve_map=True):
        """
        Load Curves from SQL Server
        
        =======================
        Parameters:
        curve_group: curve group class instance
            curve group instance to load the curves
        source_curve_group_name: int
            source curve_group_name if known
        model_name: str
            model name with curves to download
        scenario_name: str
            name of the scenario in selected model
        curve_type: str (optional)
            'all' or the name of curve type [default, prepay, curtail, rollrate, recovery]
            if all will download all curves
        curve_sub_type: str (optional)
            'all' or the name of the curve sub type [base, adjust, stress]
            if all will download all curve sub types
        """
        
        #selected_curve_group = self.rate_curve_groups[curve_group]
        
        print(('Downloading Curve Data - {} - {}').format(model_name, scenario_name))
        
        #import data
        sql_cmd = "exec fra.cf_model.usp_extract_curves ?, ?, ?, ?, ?, ?, ?"
        params = [curve_group.curve_group_key, source_curve_group_name, model_name, scenario_name, curve_type, curve_sub_type, int(update_curve_map)]

        curve_info, curve_data, segment_input, segment_map = self.sql_engine_import.execute(sql_cmd, params, output=True)
        curve_data.set_index(['curve_type', 'curve_sub_type', 'curve_id', 'from_status', 'to_status', 'period'], inplace=True)
        curve_data = curve_data.astype({'rate':'float32'})
        idx = pd.IndexSlice
                
        print('Creating Curve Data Records and Lookups')
        

        #########################################################
        #iterate over curve info and load into curve group object
        for ix, row in curve_info.iterrows():
            curve_type = row['curve_type']
            curve_sub_type = row['curve_sub_type']
            period_type = row['period_type']  
            source = row['curve_source']
            
            final_df = curve_data.loc[idx[curve_type, curve_sub_type, :, :], :]
            curve_group.add_curve(final_df, curve_type, curve_sub_type, period_type, source)
            print('    -{} {}'.format(curve_type, curve_sub_type))
            
        print('Creating Segment Data')
        
        #########################################
        #iterate over segment types and load
        for ix, row in segment_input.iterrows():

            #update segment definitions
            segment_input_dict = json.loads(row['model_input_cmd'])
            curve_group.add_segment(row['curve_type'], **segment_input_dict)
            curve_group.segments[row['curve_type']].segment_key=row['segment_key']
        
            #selected segment map
            isolated_seg_map = segment_map[segment_map['curve_type']==row['curve_type']]
            curve_map_final = {row['curve_type'] : dict(zip(isolated_seg_map['segment_name'], isolated_seg_map['curve_name']))} 
            curve_group.segment_map_manual.update(curve_map_final)
                    
    def import_rate_curves_excel(self, curve_group, curve_type, curve_sub_type, file_path, ws_name, ws_range='C3:NF363', period_type='MOB', key_cols=['Key'], key_rename=[], pivot=False):
        """
        Load Curves from source excel file
        
        =================================================================
        Parameters:
        curve_group: curve_group class instance
            curve group instance to laod curves
        curve_type: str
            type of curve to apply this segment too:
            valid input: default, prepay, curtail, recovery
            This should tie to the segment types on input data
        curve_sub_type: str
            curve sub type
            valid input: base, adjust
        file_path: raw string
            full file path. enter with "r' prefix"
        ws_name: string
            worksheet name
        ws_range: string
            cell(s) or column(s) to identify cell range
        period_type: string
            type of time period cures are organized by.
            options are: month_on_book or cal_month
        key_cols: list
            column(s) containing curve unique identifier(s)
            * this is renamed to "curve name" once loaded
        key_rename: list
            optional: rename key columns
            must be same length as key_cols
        """
        #valid_segments=['default','prepay', 'curtail', 'recovery','collection']
        if curve_type not in self.segment_types:
            raise Exception("Invalid Curve Type. Valid options are: {}".format(self.segment_types))
        
        print('Opening Worksheet: {}'.format(ws_name))
                    
        #load range
        raw_range = self.excel_engine.from_excel(file_path, ws_name, ws_range)
        #drop na cols 
        raw_range.dropna(axis=1, how='all', inplace=True)
        
        #unpivot data if loaded a horizontal table
        # turns wide table into skinny table
        if pivot:
                   
            #unpivot data set (wide table to skinny table)
            final_df = self.excel_engine.unpivot_df(raw_range, pivot_col_name='period')
               
            final_df['curve_id'] = final_df[key_cols].apply(lambda x: '|'.join(x.astype(str)), axis=1)
        
            if len(key_rename)==0:
                final_df['curve_name'] = '|'.join(key_cols)
            else:
                final_df.rename(columns = dict(zip(key_cols, key_rename)), inplace=True)
                final_df['curve_name'] = '|'.join(key_rename)
        
            #drop extra columns (anything used to create key)
            keep_cols = ['curve_id','curve_name', 'period', 'rate', 'from_status', 'to_status'] #period_type, other_period_type,
        
            #keep_cols.extend(key_cols)
            remove_cols = [x for x in final_df if x not in keep_cols]
            final_df.drop(columns=remove_cols, inplace=True)
                
        #drop na rows
        final_df.dropna(axis=0, how='any', inplace=True)
        #replace any non numeric values with zero
        final_df['rate']=(pd.to_numeric(final_df['rate'], errors='coerce')).fillna(0)
        final_df = final_df.astype({'rate':'float32'})
        
        #final_df.insert(0,'curve_source', 'excel')
        final_df.insert(0,'curve_sub_type', curve_sub_type)
        final_df.insert(0,'curve_type', curve_type)
        
        #if not roll rate add dummy data for from_status and to_status
        if curve_type == 'default':
            final_df['from_status'] = 1
            final_df['to_status'] = 8
        elif curve_type == 'prepay':
            final_df['from_status'] = 1
            final_df['to_status'] = 12
        elif curve_type != 'rollrate':
            final_df['from_status'] = -1
            final_df['to_status'] = -1
        
        final_df.set_index(['curve_id', 'period', 'from_status', 'to_status'], inplace=True) #'curve_type', 'curve_sub_type',
        
        #if this is a adjust curve Add 1. ie 20% will become 120%
        if curve_sub_type=='adjust':
            final_df['rate']= final_df['rate']+1.0
                
        curve_group.add_curve(final_df, curve_type, curve_sub_type, period_type, 'Excel Custom Curves')
              
    def export_data_sql(self, input_df, schema, table_name, if_exists='fail'):
        self.sql_engine_export.upload_df(input_df, schema, table_name, if_exists=if_exists)
        
    def create_model_config(self, model_config_name, config_type=1, config_dict={}):
        """
        Creates a new model configuration. A dict with model parameters must be passed in manually
        
        Parameters
        ====================
        model_config_name: str
            the name for the new configuration
        config_type: int
            type of model we are building
            (1: Hazard, 2: Roll Rate, 3: Monte Carlo)
        config_dict: dict
            dict containing model config parameters
        """
        
        sql_cmd='exec fra.cf_model.usp_upload_model_config ?, ?, ?'
        params=[model_config_name, config_type, json.dumps(config_dict)]
        config_id = self.sql_engine_import.execute(sql_cmd, params, output=True).iloc[0][0]
        
        return ModelConfig(config_id, model_config_name, config_type, config_dict)
        
    def download_model_config(self, model_config_name, version=None):
        """
        Downloads a model configuration by name. if no version number passed in will return most recent
        
        Parameters
        ====================
        model_config_name: str
            name of the configuration to download
        version: int
            version number (Optional)
            
        """
        if version:
            seq_filter='and seq_nbr={}'.format(version)
        else:
            seq_filter='and seq_order=1'
        
        sql_cmd="""
            select model_config_key
            	,config_name
            	,seq_nbr
            	,config_type_key
            	,config_json
            from fra.cf_model.vw_model_config
            where 1=1
            	and config_name='{}' 
            	{}
            """.format(model_config_name, seq_filter)
        
        raw_config = self.sql_engine_import.execute(sql_cmd, output=True)
        raw_config = raw_config.squeeze()
        config_dict = json.loads(raw_config['config_json'])
        
        return ModelConfig(raw_config['model_config_key'], raw_config['config_name'], int(raw_config['config_type_key']), config_dict)
        #self.model_configs[model_config_name]=new_config
        

class DataTape(object):
    """
    Class to store the monthly data and make any modifications necessary for input into model
    """
    def __init__(self, raw_tape):
        self.raw_tape = raw_tape
        self.min_date = raw_tape['AsOfDate'].min().strftime('%Y-%m-%d')
        self.max_date = raw_tape[raw_tape['BOM_PrincipalBalance']>0]['AsOfDate'].max().strftime('%Y-%m-%d')
        self.cutoff_date = self.max_date #raw_tape['AsOfDate'].max().strftime('%Y-%m-%d')
        
        self.set_cutoff_tape(self.cutoff_date)
        
        self.account_status_list = {
                    0:'deferment',
                    1:'current',
                    2:'1-29 dpd',
                    3:'30-59 dpd', 
                    4:'60-89 dpd', 
                    5:'90-119 dpd', 
                    6:'120-149 dpd', 
                    7:'150-179 dpd',
                    8:'default',
                    9:'forbearance',
                    10:'mod',
                    11:'bk',
                    12:'prepay',
                    13:'pre acquisition',
                    14:'not in repay'
                }
        self.account_status_active = np.array([1,1,1,1,1,1,1,1,0,1,1,0,0,0,1], dtype='float32')
        self.num_status = len(self.account_status_list)
        self.group_accounts = True
        self.account_grouping_cols=None
        
    def set_cutoff_tape(self, cutoff_date, account_id=None):
        """
        Isolate the data tape for month of cutoff date
        
        Parameters
        ==========================================
        cutoff_date: str
            month end date formatted as string in format 'yyyy-mm-dd'
        account_id: int
            Optional
            if provided, will limit the cutoff tape down to a single account
        """
        
        if account_id:
            self.cutoff_ix = (self.raw_tape['AsOfDate']==date.datetime.strptime(cutoff_date,"%Y-%m-%d").date()) & (self.raw_tape['BOM_PrincipalBalance']>0) & (self.raw_tape['AccountID']==account_id)
        else:
            self.cutoff_ix = (self.raw_tape['AsOfDate']==date.datetime.strptime(cutoff_date,"%Y-%m-%d").date()) & (self.raw_tape['BOM_PrincipalBalance']>0)
        self.cutoff_tape = self.raw_tape.loc[self.cutoff_ix].copy()


    def attach_curve_group(self, rate_curves, cutoff_date, group_accounts=True, account_id=None):
        """
        Attaches a curve set onto data tape and aggregates to grouping level
        
        Parameters
        ==========================================
        rate_curves: rate curves object
            the rate curves class instance to use in this model
        cutoff_date: date
            cutoff date for the data tape
        group_accounts: bool
            Optional
            if true will group accounts up to highest level possible. 
            makes model run more quickly
        account_id: int
            optional
            if provided will limit data tape down to a single account for QA purposes
        
        """
        self.rate_curves = rate_curves
        #reset cutoff tape, create account map, and join curve ids
        self.set_cutoff_tape(cutoff_date, account_id)
        #rate_curves.create_account_map()
        
        #isolate curve map from cutoff index
        cutoff_map = rate_curves.curve_account_map.loc[self.cutoff_ix.values]
        
        self.cutoff_tape = pd.merge(self.cutoff_tape.set_index('AccountKey'), cutoff_map, left_index=True, right_index=True, sort=True) #data tape input
        input_tape = self.set_account_groups(group_accounts)
        return input_tape
        
    def set_account_groups(self, group_accounts=False):
        """
        Returns a cutoff data tape at the specified grouping level.
        Option to group accounts on key static attributes and 
            calculate weighted averages on the rest.
        
        Parameters
        =======================================
        group_accounts : bool, optional
            if true, will group accounts to reduce model run time. The default is False.

        Returns
        ====================
        DataFrame
            Finalized Cutoff Data tape with curve group attached

        """
        data_tape=self.cutoff_tape.copy()
        
        if not group_accounts:
            data_tape['rec_cnt'] = 1
        
        else:
            #if grouping rules not provided create default
            if not self.account_grouping_cols:
                
                key_cols = ['DealID','BatchKey','BatchAcquisitionDate','AsOfDate','MonthsOnBook','RemainingTerm','MonthsToAcquisition',
                            'InterestRateType','InterestRateIndex','InterestRateChangeFrequency',
                            'MinInterestRate','MaxInterestRate','AccountStatusCode','PromoType', 
                            'PromoTerm', 'PromoEndMonth', 'PromoAmortTerm','PromoBalloonFlag','PromoBalloonDate', 'OriginationTerm']
                            #'FirstPaymentDueDate', 'PromoStartDate','PromoEndDate',
                sum_cols = ['OriginationBalance','PurchaseBalance','PromoTargetBalance','BOM_PrincipalBalance','InterestBalance',
                            'TotalPrincipalBalance','TotalBalance', 'InterestBearingPrincipalBalance', 'DeferredPrincipalBalance','ScheduledPaymentAmount','ContractualPrincipalPayment',
                            'InterestPayment','ScheduledPaymentMade','PrincipalPartialPrepayment', 'PrincipalFullPrepayment', 'TotalPrincipalPayment','PostChargeOffCollections',
                            'TotalPaymentMade' #MinimumPaymentDue
                        ]
                weight_avg_cols = ['InterestRate','InterestMargin','OriginationCreditScore']
                
                self.account_grouping_cols = {
                        'key_cols': key_cols,
                        'sum_cols': sum_cols,
                        'weight_avg_cols': weight_avg_cols
                        }
                
            #remove cols that are all null
            key_cols = self.account_grouping_cols['key_cols']
            for col in reversed(key_cols):
                if pd.isnull(data_tape[col]).all():
                    self.account_grouping_cols['key_cols'].remove(col)

            #add and segment colums found
            for col in data_tape.columns:
                if col in self.rate_curves.segment_types and col not in key_cols:
                    self.account_grouping_cols['key_cols'].extend([col])
                    
            #self._model_config['acct_grouping_cols']['key_cols']=key_cols
            tape_grouped = data_tape.groupby(self.account_grouping_cols['key_cols'], dropna=False).apply(self.assign_agg_functions)
            return tape_grouped.reset_index()
        
    def assign_agg_functions(self, x):
        
        agg_dict = {}
        
        #count number of loans being aggregated in each group
        agg_dict['rec_cnt'] = x.count()
        
        #sum fields
        for col in self.account_grouping_cols['sum_cols']:
            agg_dict[col] = x[col].sum()
        
        #weighted Average fields
        for col in self.account_grouping_cols['weight_avg_cols']:
            agg_dict[col] = (x[col] * x['BOM_PrincipalBalance']).sum()/x['BOM_PrincipalBalance'].sum()
        
        return pd.Series(agg_dict, index=list(agg_dict.keys()))
        
class CurveGroup(object):
    
    def __init__(self, curve_group_key, curve_group_name, data_tape, period_type='MOB'):
        
        self.curve_group_key = curve_group_key
        self.curve_group_name = curve_group_name
        self.data_tape = data_tape
        self.period_type = period_type
        self.locked = False
        
        self.segment_types=['default','prepay', 'curtail', 'recovery', 'rollrate', 'index', 'payment']
        self.curve_sub_type=['base', 'adjust', 'stress']
        
        self.curves = {}
        self.segments = {}
        
        self.curve_type_info = {}
        self.curve_keys = pd.DataFrame(columns=['curve_type', 'curve_id','curve_key']) #['curve_type', 'curve_name','curve_key']
        self.transition_keys = pd.DataFrame(columns=['curve_type', 'curve_key', 'from_status', 'to_status'])
        
        #account maps
        self.segment_map_manual = {}
        self.segment_curve_map = pd.DataFrame(columns=['segment_type','segment_name','curve_id', 'curve_key', 'segment_key'])
        self.segment_account_map = pd.DataFrame()
        self.curve_account_map = pd.DataFrame()
    
    def lock_curve_group(self):
        """
        Sets the "locked" attribute on curve sets downloaded as is. we do not want to alter prior curve
        sets. only allows alter on newly created curve sets.

        """
        self.locked = True
        
    def attach_data_tape(self, data_tape=None):
        """
        Attaches data tape to curve group and creates account maps

        Parameters
        ----------
        data_tape : Data tape object
            data tape object to attach.

        """
        
        if not data_tape:
            return
        
        self.data_tape = data_tape
        
        #run curve mapping
        self.create_account_map()
        
    def add_curve(self, curve_df, curve_type, curve_sub_type, period_type, source):
        
        if self.locked and curve_type!='index':
            raise Exception("This Curve Group is locked and cannot be modified. Please create a new Curve Group.")
        
        self.curves[(curve_type, curve_sub_type)] = CurveSet(self, curve_df, curve_type, curve_sub_type, period_type, source)
        self.update_curve_keys(curve_type)
        
        if curve_sub_type=='base':
            self.curve_type_info[curve_type] = (period_type, source)
            
        #if single curve id was loaded, just map this to all accounts
        unique_id = self.curves[(curve_type, curve_sub_type)].data_rate_curves.index.unique(level='curve_id')
        if len(unique_id)==1:
            self.add_segment(curve_type)
        
    def update_curve_keys(self, curve_type):
        
        #curve_ids=[]
        curve_ids = pd.DataFrame(columns=['curve_id'])
        transition_ids = pd.DataFrame(columns=['curve_id', 'from_status','to_status'])
        for ck, cv in self.curves.items():
            if ck[0]==curve_type:
                #curve_ids.append(cv.curve_ids['curve_id'])
                transition_ids = transition_ids.append(cv.curve_ids[['curve_id','from_status','to_status']])
                
        #add additional transition fields
        transition_ids = transition_ids.drop_duplicates()
        transition_ids.insert(0, 'curve_type', curve_type)
        transition_idx = self.transition_keys.loc[self.transition_keys['curve_type']==curve_type].index
        self.transition_keys.drop(transition_idx, inplace=True, errors='ignore')
        self.transition_keys = self.transition_keys.append(transition_ids, ignore_index=True, sort=True)
        self.transition_keys.index.rename('transition_key', inplace=True)
        
        #get unique curve ids 
        curve_keys = transition_ids[['curve_type', 'curve_id']].drop_duplicates()
        #curve_keys.insert(0, 'curve_type', curve_type)
        curve_keys.index.rename('curve_key', inplace=True)
        curve_keys.reset_index(inplace=True)
             
        #curve_keys = pd.DataFrame(pd.concat(curve_ids).unique(), columns=['curve_name'])
        #curve_keys.insert(0, 'curve_type', curve_type)
        #curve_keys.index.rename('curve_key',inplace=True)
        #curve_keys.reset_index(inplace=True)
        
        curve_idx = self.curve_keys.loc[self.curve_keys['curve_type']==curve_type].index
        self.curve_keys.drop(curve_idx, inplace=True, errors='ignore')
        self.curve_keys = self.curve_keys.append(curve_keys, ignore_index=True, sort=True)
        
        #create curve_key dictionary
        #filtered_keys = self.curve_keys[self.curve_keys['curve_type']=='default'][['curve_id', 'curve_key']]
        #curve_key_dict = pd.Series(filtered_keys['curve_id'], index=filtered_keys['curve_key']).to_dict()
        #curve_key_dict = dict(zip(self.curve_keys['curve_id'], self.curve_keys['curve_key']))
        
        #update transition frame
        key_tuples = list(self.curve_keys[['curve_type', 'curve_id']].itertuples(index=False, name=None))
        curve_key_dict = dict(zip(key_tuples, self.curve_keys['curve_key'].values))
        #self.transition_keys['curve_key'] = self.transition_keys['curve_id'].map(curve_key_dict).values
        self.transition_keys['curve_key'] = pd.Series(self.transition_keys[['curve_type','curve_id']].itertuples(index=False, name=None)).map(curve_key_dict).values
        
        curve_tuples= list(self.transition_keys[['curve_type' ,'curve_key', 'from_status', 'to_status']].itertuples(index=False, name=None))
        transition_key_dict = dict(zip(curve_tuples, self.transition_keys.index.values))
        
        #update curve key in rate curves df 
        #curve_key_dict = dict(zip(self.curve_keys['curve_name'], self.curve_keys['curve_key']))
        for k, v in self.curves.items():
            if k[0]==curve_type:
                #v.data_rate_curves['curve_key'] = v.data_rate_curves.reset_index(drop=False)['curve_id'].map(curve_key_dict).values
                v.data_rate_curves['curve_key'] = pd.Series(v.data_rate_curves.reset_index(drop=False)[['curve_type','curve_id']].itertuples(index=False, name=None)).map(curve_key_dict).values
                v.data_rate_curves['transition_key'] = pd.Series(v.data_rate_curves.reset_index(drop=False)[['curve_type' ,'curve_key','from_status','to_status']].itertuples(index=False, name=None)).map(transition_key_dict).values
        
                
    def add_segment(self, segment_type, **kwargs):
        
        if self.locked and segment_type!='index':
            raise Exception("This Curve Group is locked and cannot be modified. Please create a new Curve Group.")
        
        self.segments[segment_type] = SegmentSet(self, segment_type, **kwargs)
        
    def map_segments_to_curves(self, segment_type):
        """
        Maps rate curves to segments by Key column names. matching values will automatically be mapped. 
        Any unmatched segment names will need to be manually mapped.
        
        match occurs on 'rule_name' column and 'curve_id' column
        =======================
        Parameters:
        segment_type: str
            type of segment ()
        manual_map: dataframe
            optional manual mapping df
        """

        curve_found = False
        segment_found = False

        #validate curves exist
        for k in self.curves:
            if k[0]==segment_type:
                curve_found=True
        
        if not curve_found:
            print(f"No curves exist yet for {segment_type}. Segment/Curve map was not created.")
            return
            
        #validate Segments exists
        for k in self.segments:
            if k==segment_type:
                segment_found=True
        
        if not segment_found:
            print(f"No segments exist yet for {segment_type}. Segment/Curve map was not created.")
            return
        
        unique_curves = self.curve_keys[self.curve_keys['curve_type']==segment_type]['curve_id'] #curve_name
        unique_segments = self.segments[segment_type].segment_rules_combined['rule_name_combined']
        
        #generate dict to map full unique values
        segment_curve_map = dict.fromkeys(unique_segments.to_list())
        for x in unique_curves:
            if x in segment_curve_map.keys():
                segment_curve_map[x]=x
        
        #if manual map exists overwrite values 
        manual_map = self.segment_map_manual.get(segment_type)
        
        if manual_map:
            try:
                segment_curve_map = {**segment_curve_map, **manual_map}
            except:
                pass
                
        #check for missing curves
        missing_curves={k:v for k,v in segment_curve_map.items() if v is None}
        
        if len(missing_curves)>0:
            missing_segments = ', '.join('"{0}"'.format(s) for s in missing_curves.keys())
            print(f"At least one {segment_type} segment was not matched to a curve. Manually map curve names with function 'map_curves()' \n\n Unmapped Segments: {missing_segments}")
        elif len(unique_segments)==0:
            pass
        else:
            print(f"All {segment_type} segments matched successfully")
        
        output_df = pd.DataFrame(list(segment_curve_map.items()), columns=['segment_name', 'curve_id']) #curve_name
        output_df.insert(0, 'segment_type', segment_type)
        #map curve keys
        curve_key_selection = self.curve_keys[self.curve_keys['curve_type']==segment_type]
        curve_key_dict = dict(zip(curve_key_selection['curve_id'], curve_key_selection['curve_key'])) #curve_name

        output_df['curve_key'] = output_df['curve_id'].map(curve_key_dict) #curve_name
        #map segment keys
        segment_keys = self.segments[segment_type].segment_rules_combined['rule_name_combined'].reset_index().set_index('rule_name_combined')
        output_df['segment_key'] = output_df['segment_name'].map(segment_keys['index'])
        
        #delete exising entries and load new map
        curve_index = self.segment_curve_map[self.segment_curve_map['segment_type']==segment_type].index
        self.segment_curve_map.drop(curve_index, inplace=True, errors='ignore')
        self.segment_curve_map = self.segment_curve_map.append(output_df, ignore_index=True)
            
    
    def return_transition_matrix_old(self):
        """
        Generates transition matrix from CDR/CPR matrix inputs.
        Model calculations are roll rate driven so we need to convert these two curves 
        into a transition matrix. Essentially, CDR is converted to the Current to Default roll 
        and CPR is converted to the Current to Prepay roll. 
        """
        
        #isolate CDR/CPR cohorts and create distinct groups
        rr_segments = self.curve_account_map[['rollrate','default','prepay']].copy()
        rr_segments['MonthsOnBook'] = self.segment_account_map['MonthsOnBook']
        rr_segments.reset_index(inplace=True, drop=True)
        rr_segments.drop_duplicates(inplace=True)
        rr_segments = rr_segments.set_index('rollrate').sort_index()
        mob = rr_segments.pop('MonthsOnBook')
        
        #rename columns to ID status number
        #swap account status desc with keys and rename columns
        account_status_list_swap = {v:k for k,v in self.data_tape.account_status_list.items()}
        rr_segments.rename(columns=account_status_list_swap, inplace=True)
        
        #create all other status columns ## CDR=8, CPR=12, added columns will be NaN
        rr_segments.columns = rr_segments.columns.astype(str)
        rr_segments = rr_segments.assign(**{str(col):(rr_segments[str(col)] if str(col) in rr_segments.columns.values else 'empty') for col in list(self.data_tape.account_status_list.keys())})
        rr_segments.columns = rr_segments.columns.astype(int)
        rr_segments.sort_index(axis=1, inplace=True)
        
        #map rate curves onto account list
        segments_array = self.map_rate_arrays(rr_segments, self.return_curve_group(['default','prepay']), output_type=1, mob=mob)
        return segments_array
        #return rr_segments, mob
        
    def return_transition_matrix(self):
        """
        Generates a 4D transition matrix from 1D input lists. 

        Returns
        -------
        4D numpy array

        """
        ##NEED TO START WITH ALL POSSIBLE CURVE KEYS FROM THE CURVE_KEYS FRAME. ANYTHING LESS WILL BREAK THE ARRAYS WHEN THEY ARE CREATED
        ##USING THE ROLLRATE KEY CREATED IN CURVE ACCOUNT MAP WILL WORK ON CDR/CPR MODELS
        ## BUT WILL BREAK IF ACTUALL ROLL RATES ARE UPLOADED AND ONE OF THE SEGMENTS IS NOT USED
        ## PROBABLY NEED TO ADD IN THE DISTINCT ROLLRATE KEYS INTO THE CURVE_KEY FRAME
        
        ##### isolate cohorts and create distinct groups #####
        #get distinct roll rate curve ids
        rr_keys = self.curve_keys[self.curve_keys['curve_type']=='rollrate'][['curve_key', 'curve_type']].copy()
        
        #get account map
        #rr_segments = self.curve_account_map[['rollrate', 'default', 'prepay']].loc[self.data_tape.cutoff_ix.values].copy()
        rr_segments = self.curve_account_map.loc[self.data_tape.cutoff_ix.values, self.curve_account_map.columns.isin(['rollrate','default','prepay'])].copy()
        
        
        #rr_segments['MonthsOnBook'] = self.segment_account_map['MonthsOnBook'].loc[self.data_tape.cutoff_ix.values]
        rr_segments.reset_index(inplace=True, drop=True)
        rr_segments.drop_duplicates(inplace=True)
        #rr_segments = rr_segments.set_index('rollrate').sort_index()
        #mob = rr_segments.pop('MonthsOnBook')
        
        #if CDR/CPR unpivot keys
        if all(x in rr_segments.columns for x in ['default', 'prepay']):
            #unpivot keys
            rr_segments = pd.melt(rr_segments, id_vars=['rollrate'], var_name='curve_type', value_name='curve_key', ignore_index=False)
            #merge with rate curves
            rr_segments = rr_segments.merge(self.transition_keys.reset_index()[['transition_key','curve_type','curve_key','from_status', 'to_status']], how='left', left_on=['curve_type', 'curve_key'], right_on=['curve_type', 'curve_key'])
            rr_segments = rr_segments.set_index('rollrate').sort_index()[['transition_key', 'from_status', 'to_status']]
            #duplicate default segments across all dpd statuses
            default_ix = (rr_segments['from_status']==1) & (rr_segments['to_status']==8)
            rr_duplicate_d = pd.concat([rr_segments[default_ix.values]]*6) # , ignore_index=True)
            rr_duplicate_d['from_status'] = rr_duplicate_d.groupby(['rollrate']).cumcount()+2
            rr_segments = pd.concat([rr_segments, rr_duplicate_d])
            #duplicate prepay segments across all dpd statuses
            prepay_ix = (rr_segments['from_status']==1) & (rr_segments['to_status']==12)
            rr_duplicate_p = pd.concat([rr_segments[prepay_ix.values]]*6)
            rr_duplicate_p['from_status'] = rr_duplicate_p.groupby(['rollrate']).cumcount()+2
            rr_segments = pd.concat([rr_segments, rr_duplicate_p])
            
        else:
            rr_segments = rr_keys
            rr_segments = rr_segments.merge(self.transition_keys.reset_index()[['transition_key','curve_type','curve_key','from_status', 'to_status']], how='left', left_on=['curve_type', 'curve_key'], right_on=['curve_type', 'curve_key'])
            rr_segments = rr_segments.set_index('rollrate').sort_index()[['transition_key', 'from_status', 'to_status']]
            
        #rr_segments = rr_segments.merge(self.transition_keys.reset_index()[['transition_key','curve_type','curve_key','from_status', 'to_status']], how='left', left_on=['curve_type', 'curve_key'], right_on=['curve_type', 'curve_key'])
        #rr_segments = rr_segments.set_index('rollrate').sort_index()[['transition_key', 'from_status', 'to_status']]
                
        #get curve array
        rate_data = self.return_curve_group(['default','prepay','rollrate'])
        
        #generate rate mapping dictionary
        rate_dict = self.map_rate_arrays(rr_segments, rate_data.droplevel(['curve_type', 'curve_key']))
        rr_segments['rate'] = rr_segments['transition_key'].map(rate_dict)
        
        #create empty matrix
        #rr_matrix = np.full(shape=[len(rr_segments), self.data_tape.num_status, self.data_tape.num_status], fill_value=-1, dtype=int)
        rr_matrix = np.zeros(shape=[len(rr_segments), self.data_tape.num_status, self.data_tape.num_status, len(rate_dict[-1])], dtype=np.float32)
        #populate matrix with transition ids
        rr_segment_index = rr_segments.reset_index(drop=False)[['rollrate', 'from_status','to_status']].values.astype(int)
        rr_matrix[rr_segment_index[:,0], rr_segment_index[:,1], rr_segment_index[:,2]] = np.array(rr_segments['rate'].to_list())
        
        
        #swap axis to final configuration
        #current (curvekey x from_status x to_status x time)
        #final   (curvekey x time x from_status x to_status)
        array_reshape = np.moveaxis(rr_matrix, 3,1)
        
        #create account status columns
        return array_reshape
        #return rr_segments
        
    def return_rate_matrix(self, segment_type):
        """
        converts a set of segments curves into a lookup array for input into model
        
        =====================
        Parameters
        segment_type: str
            segment type to return. valid options are stored in the segment_type attribute
        """
        #isolate distinct segment ids
        #rate_segments = self.curve_account_map[segment_type].copy()
        #rate_segments = self.curve_account_map.loc[self.data_tape.cutoff_ix.values, [segment_type]].copy()
        
        ##NEED TO START WITH ALL POSSIBLE CURVE KEYS FROM THE CURVE_KEYS FRAME. ANYTHING LESS WILL BREAK THE ARRAYS WHEN THEY ARE CREATED
        
        rate_segments = self.curve_keys[self.curve_keys['curve_type']==segment_type][['curve_key', 'curve_type']].copy()
        
        #rate_segments.reset_index(inplace=True, drop=True)
        #rate_segments.drop_duplicates(inplace=True)
        #rate_segments['curve_type'] = segment_type
        #rate_segments['curve_key'] = rate_segments[segment_type]
        #rate_segments = rate_segments.sort_values().reset_index(drop=True)
        
        rate_segments = rate_segments.merge(self.transition_keys.reset_index()[['transition_key','curve_type','curve_key','from_status', 'to_status']], how='left', left_on=['curve_type', 'curve_key'], right_on=['curve_type', 'curve_key'])
        rate_segments = rate_segments.set_index('curve_type').sort_index()[['transition_key', 'from_status', 'to_status']]
        
        
        #segment_array = self.map_rate_arrays(rate_segments, self.return_curve_group([segment_type]), output_type=2)
        #get curve array
        rate_data = self.return_curve_group([segment_type])
        
        #generate mapping dict
        rate_dict = self.map_rate_arrays(rate_segments, rate_data.droplevel(['curve_type','curve_key']))
        rate_segments['rate'] = rate_segments['transition_key'].map(rate_dict)
        
        #extract final matrix
        rate_matrix = np.array(rate_segments['rate'].to_list())
        
        return rate_matrix
        #return rate_segments
    
    def return_curve_group(self, curve_types=[]):
        """
        Applies adjust curves as a multiplier then returns final curve set for input to model.
        Groups by all columns except for rate and applies the Prod() function. 
        If no adjust curve is present, then there is no change
        if a adjust curve is present, it is multiplied onto the base curve
        """
        
        curve_dfs = []
        
        for k, v in self.curves.items():
            if k[0] in curve_types:
                curve_dfs.append(v.data_rate_curves.reset_index())
        
        curve_df = pd.concat(curve_dfs, sort=True)
        #align data types
        curve_df = curve_df.astype({'curve_type':'str', 'curve_id':'str', 'period':'int32', 'curve_key':'int32', 'from_status':'int32', 'to_status':'int32', 'rate':'float32'})
        #curve_df = curve_df.set_index(['curve_type','curve_id','period','curve_key','from_status','to_status'])['rate']
        curve_df = curve_df.set_index(['curve_type','curve_id','curve_key', 'transition_key','period'])['rate']
        
        curve_group_final = curve_df.groupby(curve_df.index.names, sort=False).prod()
        #return curve_group_final.reset_index().set_index(['curve_type', 'curve_key','period']).sort_index()['rate'] #rate curves input    #'segment_id'
        #return curve_group_final.reset_index().set_index(['curve_type', 'curve_key','from_status','to_status', 'period']).sort_index()['rate'] #rate curves input    #'segment_id'
        return curve_group_final.reset_index().set_index(['curve_type', 'curve_key', 'transition_key', 'period']).sort_index()['rate'] #rate curves input    #'segment_id'
        #return curve_df
        
    def return_historical_defaults(self):
        defaults = self.data_tape.raw_tape[['AccountKey', 'BatchKey', 'AsOfDate', 'ChargeOffAmount', 'PostChargeOffCollections']].copy()
        #match default segments
        for ix, r in self.segments['recovery'].segment_rules_combined['rule_eval_combined'].iteritems():
            #self.segment_account_map.loc[r[self.data_tape.cutoff_ix], 'segment_'+segment_type] = ix
            defaults.loc[r, 'recovery_segment'] = ix
        #match curves
        selected_map = self.segment_curve_map[self.segment_curve_map['segment_type']=='recovery'].copy()
        selected_map.dropna(how='any', inplace=True)
        curve_key_dict = dict(zip(selected_map['segment_key'], selected_map['curve_key']))
        defaults['recovery_curve'] = defaults['recovery_segment'].map(curve_key_dict)
        
        defaults = defaults[['BatchKey','AsOfDate','recovery_curve','ChargeOffAmount']].groupby(['BatchKey', 'recovery_curve', 'AsOfDate']).sum()
        #self.data_tape.cutoff_date
        defaults['cutoff_date'] = date.datetime.strptime(self.data_tape.cutoff_date,"%Y-%m-%d").date()
        defaults.reset_index(inplace=True)
        defaults['ProjectionMonth'] = defaults.apply(lambda row: self.time_dif_months(row['cutoff_date'], row['AsOfDate']), axis=1) * -1
        
        return defaults[['BatchKey','AsOfDate','recovery_curve','ProjectionMonth','ChargeOffAmount']].set_index(['AsOfDate', 'BatchKey','recovery_curve']).sort_index()
        
    def time_dif_months(self, date_1, date_2):
        
        time_diff = relativedelta.relativedelta(date_1, date_2)
        time_diff = time_diff.years*12 + time_diff.months
        return time_diff
    
    
    def create_account_map(self):
        
        #add months on book just in case we have a mix of MonthonBook curves and Calendar month curves 
        #self.segment_account_map = self.data_tape.cutoff_tape[['AccountKey','BatchKey','MonthsOnBook']].drop_duplicates().set_index('AccountKey')
        #self.curve_account_map = pd.DataFrame(index=self.segment_account_map.index)
        
        #self.segment_account_map = self.data_tape.raw_tape[['AccountKey','AsOfDate','BatchKey','MonthsOnBook']].drop_duplicates().set_index('AccountKey', 'AsOfDate')
        self.segment_account_map = self.data_tape.raw_tape[['AccountKey','AsOfDate','BatchKey']].drop_duplicates().set_index('AccountKey', 'AsOfDate')
        self.curve_account_map = pd.DataFrame(index=self.segment_account_map.index)
        
        #for segment in self.segment_types:
        for segment in self.segments.keys():
            self.match_segment(segment)
            self.match_curve(segment)
            self.update_segment_group_id()
        #update RR ID if both Default and Prepay curves exist
        self.update_curve_group_id()
    
    
    def match_segment(self, segment_type):
        """
        Add created segment onto the account map as a new column
        
        =======================
        Parameters:
        segment_type: str
            type of segment
        """
                 
        for ix, r in self.segments[segment_type].segment_rules_combined['rule_eval_combined'].iteritems():
            #self.segment_account_map.loc[r[self.data_tape.cutoff_ix], 'segment_'+segment_type] = ix
            #self.segment_account_map.loc[r, 'segment_'+segment_type] = ix
            self.segment_account_map.loc[r, segment_type] = ix
            
    def match_curve(self, curve_type):
        """
        map each account from segment id to curve id
        """
        #if segment type does not exist then do nothing
        #if 'segment_'+curve_type in self.segment_account_map.columns:
        if curve_type in self.segment_account_map.columns:
            #map ids to accounts
            #curve_key_dict = self.segments[curve_type].segment_rules_combined['curve_key'].to_dict()
            selected_map = self.segment_curve_map[self.segment_curve_map['segment_type']==curve_type].copy()
            selected_map.dropna(how='any', inplace=True)
            curve_key_dict = dict(zip(selected_map['segment_key'], selected_map['curve_key']))
            #self.curve_account_map[curve_type] = self.segment_account_map['segment_'+curve_type].map(curve_key_dict)
            self.curve_account_map[curve_type] = self.segment_account_map[curve_type].map(curve_key_dict)
            
    def update_segment_group_id(self):
        """
        Creates a unique id for each distinct combination of segments
        """
        grouping_cols = ['BatchKey']
        segment_cols = [col for col in self.segment_account_map.columns.values]
        grouping_cols.extend(segment_cols)
        
        self.segment_account_map['segment_group_id']=self.segment_account_map.groupby(grouping_cols).ngroup()        
    
    def update_curve_group_id(self):
        #create RR id if both default and prepay curves exist
        if all(x in self.curve_account_map.columns for x in ['default','prepay']):
            #self.curve_account_map['rollrate'] = self.curve_account_map.groupby(['MonthsOnBook','default','prepay']).ngroup()
            #add to curve account map
            self.curve_account_map['rollrate'] = self.segment_account_map.groupby(['default', 'prepay']).ngroup() #['MonthsOnBook', 'segment_default', 'segment_prepay']
            
            #add to curve_keys
            rr_unique = self.curve_account_map[['rollrate']].drop_duplicates()
            rr_unique.insert(0, 'curve_id', 'rr_' + rr_unique['rollrate'].astype(str))
            rr_unique['curve_type'] = 'rollrate'
            rr_unique.rename(columns={'rollrate':'curve_key'}, inplace=True)
            rr_unique.sort_values(by='curve_key', inplace=True)
            rr_unique.reset_index(drop=True, inplace=True)
                        
            curve_idx = self.curve_keys.loc[self.curve_keys['curve_type']=='rollrate'].index
            self.curve_keys.drop(curve_idx, inplace=True, errors='ignore')
            self.curve_keys = self.curve_keys.append(rr_unique, ignore_index=True, sort=True)
            
    def update_all_mappings(self, manual_map=None):
        if manual_map:
            self.segment_map_manual.update(manual_map)
        
        for segment in self.segment_types:
            self.map_segments_to_curves(segment)
    
    def map_rate_arrays(self, account_segments, rate_data):
        """
        Takes curve keys ID by account and maps in entire rate curve onto that distinct account group
        
        =======================
        Parameters:
        account_segments: dataframe
            dataframe with accoumt ids as key and curve_keys as columns
        mob: dataframe/series
            dataframe column containing MOB by             
        rate_data: dataframe
            dataframe with monthly rates by segment id
        output_type: int (1 or 2)
            1 = Transition matrix, returns a 13x13 matrix for each account, each month
            2 = single value, returns a single value for each account, each month
        """
        max_length = max(rate_data.groupby(['transition_key']).count())
        max_rem_term = max(self.data_tape.cutoff_tape['RemainingTerm'])
        max_mob = max(self.data_tape.cutoff_tape['MonthsOnBook'])
        
        #find maximum length to extend rate arrays. 
        #either existing max length array or mob+rem_term 
        self.max_mob = int(max(max_length, (max_rem_term+max_mob+1)))
        
        self.neg_mob = min(self.data_tape.cutoff_tape['MonthsOnBook'])
        self.neg_mob = min(self.neg_mob, 0) #neg MOB must be <=0
        
        #if mob is None:
        #    mob = np.zeros(len(account_segments))
        
        #pad tail zeros
        rate_data_fill = rate_data.groupby(['transition_key']).apply(self.fill_missing_vintage)
        #add tail records for neg MOB (negative index will go to end of array)
        rate_data_fill = rate_data_fill.groupby(['transition_key']).apply(self.fill_neg_mob)
                
        rate_dict = rate_data_fill.groupby('transition_key').apply(list).to_dict()
        zeros = [0]*(self.max_mob-self.neg_mob)
        rate_dict[-1] = zeros
                
        #add zero list
        #return account_segments
        return rate_dict
        
    """
        #map curves and reshape final output
        if output_type==1: #several rate curve inputs
            #convert curves to a dict of lists
            #rate_dict = {l: rate_data_fill.xs(l, level=0).groupby('transition_key').apply(list).to_dict() for l in rate_data_fill.index.levels[0]}
            
        
            #account_segments[8] = account_segments[8].map(rate_dict['default'])
            #account_segments[12] = account_segments[12].map(rate_dict['prepay'])
        
            #fill other columns with zeros list
            for col in account_segments.columns:
                if col not in [8,12]:
                    account_segments[col] = account_segments[col].map(empty_status_dict)
            
            #roll arrays if period type is MOB
            account_segments['MonthsOnBook'] = mob
            
            if self.curve_type_info['default'][0]=='MOB':
                account_segments[8] = account_segments.apply(lambda row: np.roll(row[8], -row['MonthsOnBook']), axis=1)
            if self.curve_type_info['prepay'][0]=='MOB':
                account_segments[12] = account_segments.apply(lambda row: np.roll(row[12], -row['MonthsOnBook']), axis=1)
            
            account_segments.drop('MonthsOnBook', axis=1, inplace=True)
            
            #flatten account segment list and reshape to nd array
            account_segments_flat = account_segments.to_numpy().flatten()
            #convert from object array. fastest option to create zero array and then fill
            output_shape=(len(account_segments_flat), self.data_tape.num_status, (self.max_mob-self.neg_mob))
            segments_array = np.zeros(output_shape, dtype='float32')
            for i, v in enumerate(account_segments_flat):
                segments_array[i] = v
            #reshape to final output
            array_reshape = segments_array.reshape(len(account_segments), self.data_tape.num_status, self.data_tape.num_status, (self.max_mob-self.neg_mob))
            array_reshape = np.moveaxis(array_reshape, 3,1)
            array_reshape = array_reshape.swapaxes(2,3)
            
        elif output_type==2: #one rate curve input
            #convert DF into 2d array (curve_key x MOB)
            array_reshape = rate_data_fill.values.reshape(int(rate_data_fill.index.max()[1]+1), (self.max_mob-self.neg_mob))
        
        return array_reshape.astype('float32')
    """
    
    def map_rate_arrays_old(self, account_segments, rate_data, output_type, mob=None):
        """
        Takes segment ID by account and maps in entire rate curve onto that distinct account group
        
        =======================
        Parameters:
        account_segments: dataframe
            dataframe with accoumt ids as key and curve_keys as columns
        mob: dataframe/series
            dataframe column containing MOB by             
        rate_data: dataframe
            dataframe with monthly rates by segment id
        output_type: int (1 or 2)
            1 = Transition matrix, returns a 13x13 matrix for each account, each month
            2 = single value, returns a single value for each account, each month
        """
        max_length = max(rate_data.groupby(['curve_type','curve_key']).count())
        max_rem_term = max(self.data_tape.cutoff_tape['RemainingTerm'])
        max_mob = max(self.data_tape.cutoff_tape['MonthsOnBook'])
        
        #find maximum length to extend rate arrays. 
        #either existing max length array or mob+rem_term 
        self.max_mob = int(max(max_length, (max_rem_term+max_mob+1)))
        
        self.neg_mob = min(self.data_tape.cutoff_tape['MonthsOnBook'])
        self.neg_mob = min(self.neg_mob, 0) #neg MOB must be <=0
        
        if mob is None:
            mob = np.zeros(len(account_segments))
        
        #pad tail zeros
        rate_data_fill = rate_data.groupby(['curve_type','curve_key']).apply(self.fill_missing_vintage)
        #add tail records for neg MOB (negative index will go to end of array)
        rate_data_fill = rate_data_fill.groupby(['curve_type','curve_key']).apply(self.fill_neg_mob)
                
        #map curves and reshape final output
        if output_type==1: #several rate curve inputs
            #convert curves to a dict of lists
            rate_dict = {l: rate_data_fill.xs(l, level=0).groupby('curve_key').apply(list).to_dict() for l in rate_data_fill.index.levels[0]}
            zeros = [0]*(self.max_mob-self.neg_mob)
            empty_status_dict = {'empty':zeros}
        
            account_segments[8] = account_segments[8].map(rate_dict['default'])
            account_segments[12] = account_segments[12].map(rate_dict['prepay'])
        
            #fill other columns with zeros list
            for col in account_segments.columns:
                if col not in [8,12]:
                    account_segments[col] = account_segments[col].map(empty_status_dict)
            
            #roll arrays if period type is MOB
            account_segments['MonthsOnBook'] = mob
            
            if self.curve_type_info['default'][0]=='MOB':
                account_segments[8] = account_segments.apply(lambda row: np.roll(row[8], -row['MonthsOnBook']), axis=1)
            if self.curve_type_info['prepay'][0]=='MOB':
                account_segments[12] = account_segments.apply(lambda row: np.roll(row[12], -row['MonthsOnBook']), axis=1)
            
            account_segments.drop('MonthsOnBook', axis=1, inplace=True)
            
            #flatten account segment list and reshape to nd array
            account_segments_flat = account_segments.to_numpy().flatten()
            #convert from object array. fastest option to create zero array and then fill
            output_shape=(len(account_segments_flat), self.data_tape.num_status, (self.max_mob-self.neg_mob))
            segments_array = np.zeros(output_shape, dtype='float32')
            for i, v in enumerate(account_segments_flat):
                segments_array[i] = v
            #reshape to final output
            array_reshape = segments_array.reshape(len(account_segments), self.data_tape.num_status, self.data_tape.num_status, (self.max_mob-self.neg_mob))
            array_reshape = np.moveaxis(array_reshape, 3,1)
            array_reshape = array_reshape.swapaxes(2,3)
            
        elif output_type==2: #one rate curve input
            #convert DF into 2d array (curve_key x MOB)
            array_reshape = rate_data_fill.values.reshape(int(rate_data_fill.index.max()[1]+1), (self.max_mob-self.neg_mob))
        
        return array_reshape.astype('float32')
        
    def fill_missing_vintage(self, x):
        """
        All Rate Curves must all be the same length. ie a 120 month term curve may only extend 120 months 
        but if there are 240 loans the model will break after we get past month 240
        This function extends each array to the max length and does a "forward fill" to continue 
        the last number forward. This extends the arrays by filling the last value available
        """
        #idx = range(self.max_mob_range)
        idx = range(self.max_mob)
        #x.reset_index(level=['curve_type','curve_key'], drop=True, inplace=True)
        x.reset_index(level=['transition_key'], drop=True, inplace=True)
        return x.reindex(idx, method='ffill')
    
    def fill_neg_mob(self, x):
        """
        when backtesting FF deals there will be negative MOB until the asofdate gets to the batch 
        acquisition date. The model reads Negative MOB as a negative index. so a MOB of -2 will lookup up
        rates from the second to last item in the array. 
        
        to account for this behavior this function adds additional records to the end of the array. 
        these records will have a value of "0". that way when the transition matrix gets generated 
        all balance will stay in the same status. when the model advances to MOB=0 the lookups will
        function normally.
        """
        idx = range(self.max_mob-self.neg_mob)
        
        #x.reset_index(level=['curve_type', 'curve_key'], drop=True, inplace=True)
        x.reset_index(level=['transition_key'], drop=True, inplace=True)
        return x.reindex(idx, fill_value=0)
    
class CurveSet(object):
    
    def __init__(self, curve_group, curve_df, curve_type, curve_sub_type='base', period_type='MOB', source='Excel Custom Curves'):
        """
        Imports curve and stores descriptive info. 
        
        Parameters
        ===========================
        curve_group: parent curve group instance
            reference to the parent class curve group
        curve_df: DataFrame
            dataframe containing rate curve data
        curve_type: string
            type of curve being loaded 
            'default', 'prepay','curtail','recovery','rollrate'
        curve_sub_type: string
            curve sub type
            'base','adjust','stress'
        period_type: string
            is the curve based on month on book or calendar month
            'MOB', 'calendar'
        """
        self.curve_group = curve_group
        self.curve_type = curve_type
        self.curve_sub_type = curve_sub_type
        
        #curve info
        self.curve_type_info = pd.DataFrame(columns=['curve_type', 'curve_sub_type','period_type','source'])
        self.curve_ids = pd.DataFrame(columns=['curve_id', 'from_status', 'to_status'])
        self.data_rate_curves = pd.DataFrame()
        
        #store curve information
        self.curve_type_info = {'curve_type':curve_type, 'curve_sub_type':curve_sub_type , 'period_type':period_type , 'source':source}
       
        #store curve ids
        #self.curve_ids = curve_df.reset_index()[['curve_id']].drop_duplicates().sort_values('curve_id')
        self.curve_ids = curve_df.reset_index()[['curve_id', 'from_status', 'to_status']].drop_duplicates()
        self.curve_ids.reset_index(drop=True, inplace=True)
        
        #load curves into df
        self.data_rate_curves = curve_df.sort_index()
        
    def curve_json(self):
        upload_df = self.data_rate_curves.reset_index()[['curve_id', 'period', 'rate','from_status','to_status']]
        return upload_df.to_json(orient='records')

    def prep_curves_for_upload(self):
        upload_df = self.data_rate_curves.reset_index()[['curve_id', 'period', 'rate','from_status','to_status']]
        return upload_df.to_json(orient='records')    
        
class SegmentSet(object):
    
    def __init__(self, curve_group, segment_type, **kwargs):
        """
        Segment set object. creates segments based on kwargs input
        
        Parameters
        ==============================================
        curve_group: parent curve group instance
            reference to the parent class curve group
        curve_type: string
            type of segment being created 
            'default', 'prepay','curtail','recovery','rollrate'
        **kwargs : field_name=[]
            any number of input fields with list arg. 
            list is the breakpoints to use. if no breakpoints needed, use an empty list
            empty list will use each unique value in field 

        """
        self.curve_group = curve_group
        #initialize fields
        self.segment_type = segment_type
        #self.data_tape = data_tape        
        self.segment_input = {}
        self.segment_rules = pd.DataFrame()
        self.segment_rules_combined = pd.DataFrame(columns=['rule_name_combined','column_combined','rule_eval_combined'])
        self.segment_key = None
        self.segment_account_map = pd.DataFrame()
        
        self.create_segment(**kwargs)
    
    def create_segment(self, **kwargs):
        """
        creates segments on input data tape from specified column names. can create segments for several different types of curves 
        Pass in any number of column names with optional breakpoints if desired. 
        
        Parameters
        =================
        kwargs: list of column names
            list of column names with optional breakpoints
    
            ie. CurrentCreditScore=[600,700,800] --- will create FICO buckets <600, 600-700, 700-800 and >=800
                OriginationTerm = [] --- will use each unique value in the specified column
        """
        
        self.segment_input=copy.deepcopy(kwargs)
                        
        combined_rules = []
        
        #if data tape does not exist yet, end function early
        if not self.curve_group.data_tape:
            print(f'Data Tape does not exist yet. {self.segment_type} Segments not generated.')
            return
            
        print("""=================================================\nGenerating {} segments: 
                """.format(self.segment_type))
                
        #############################################
        #generate match rules from input columns
        
        #if no kwargs entered, create 1 global segment 
        if len(kwargs)==0:
            rule_eval = np.ones(len(self.curve_group.data_tape.raw_tape), dtype=bool)
            combined_rules.append({"column":'none', "column_order":1, "rule_name":'all', "rule_eval":rule_eval, "min_value": None, "max_value":None})
        else:
            i=1
            for key, value in kwargs.items():
                print("column={} : breakpoints={}".format(key, value))
                column_name = key
                values = value
                
                if column_name not in self.curve_group.data_tape.raw_tape.columns:
                    raise Exception(("{} is not a valid column name. Check data_tape column names").format(key))
        
                values.sort()
                
                #create bins
                if len(values) > 0:
                    column_values, self.break_points, bin_names, bin_range = self.create_bins(self.curve_group.data_tape.raw_tape[column_name], values) 
                    column_name = column_name + '_bin'
                    bin_range_map = dict(zip(bin_names, bin_range))
                else: 
                    column_values = self.curve_group.data_tape.raw_tape[column_name]
                    
                #use distinct values in column, if bins were created this is the unique bin values, if no bins created this is each unique value in the original field
                unique_values = column_values.unique()
                for val in unique_values:
                    rule_name = val
                    rule_eval = column_values==val
                    if len(values)==0:
                        bin_min_max = (val,val)
                    else:
                        #map the bin_ranges
                        try:
                            bin_min_max = bin_range_map[val]
                        except:
                            #should only trigger on Nan inputs
                            bin_min_max = (None,None)
            
                    combined_rules.append({"column":column_name, "column_order":i, "rule_name":rule_name, "rule_eval":rule_eval, "min_value": bin_min_max[0], "max_value":bin_min_max[1]})
                
                i+=1
        
        print("=================================================\n")
        
        #add rules onto master list to generate unique index
        rules = pd.DataFrame(combined_rules)
        
        self.segment_rules = rules
        
        #######################################################################
        # Combine individual rules together to get final combined segment rule
        column_list = rules['column'].unique()
        cross_dict = {}
        
        #cartesian product input lists to get each unique combination
        for col in column_list:
            ix_list = (rules[rules['column']==col].index).tolist()
            cross_dict[col+'_key']=ix_list
    
        #cross join the distinct keys from each column
        cross_df = pd.DataFrame(dict(zip(cross_dict, x)) for x in itertools.product(*cross_dict.values()))
    
        for col in column_list:
            cross_df = cross_df.merge(rules[['column', 'rule_name', 'rule_eval']], left_on=col+'_key', right_index=True)
        #cross_df.insert(0,'segment_type',segment_type)
        
        #create combined columns from individual rules
        self.combine_segment_rules(cross_df, 'rule_name')
        self.combine_segment_rules(cross_df, 'column')
        
        #combine rule_eval columns
        rule_name_cols = [col for col in cross_df.columns if 'rule_eval' in col]
        # separate the truth arrays, reduce and apply np logical to each row along axis 1 to compare each value
        cross_df['rule_eval_combined'] = cross_df[rule_name_cols].apply(lambda row: np.logical_and.reduce([i for i in row.values]), axis=1)
        cross_df.drop(columns=rule_name_cols, inplace=True)
        
        #add rules onto master list to generate unique index
        cross_df.index = cross_df.index + int(np.nan_to_num(self.segment_rules_combined.index.max()))+1
        self.segment_rules_combined = pd.concat([self.segment_rules_combined, cross_df])
        self.segment_rules_combined.sort_index(inplace=True)
        #cross_df = self.segment_rules_combined[self.segment_rules_combined['segment_type']==segment_type]
    
    
    def create_bins(self, column, break_points):
        """
        creates bins from input columns
        bins are all right inclusive
        so breakpoint [700] creates two bins:
            1) 0 - 700 (right value is included)
            2) 701-inf
            
        Parameters
        =========================
        column: pd.Series
            series for column we want to create bins for
        break_points: list
            if empty will use all unique values, if exists will use as breakpoints on Bins
            bins will be right inclusive i.e. breakpoints=[700] will return 2 bins, [-inf,700][701,inf]
        """      
        
        bin_names = []
        bin_range = []
    
        col_min = column.min()
        col_max = column.max()
    
        for x in range(len(break_points)):
            #first item
            if x == 0 and break_points[x]>col_min:
                bin_names.append("<" + str(break_points[x]))
                bin_range.append((None,break_points[x]))
            elif x == 0:
                pass
            #middle items
            else:
                bin_names.append(str(break_points[x-1])+"-"+(str(break_points[x])))
                bin_range.append((break_points[x-1],break_points[x]))
            #last item
            if x == len(break_points)-1 and break_points[x]<col_max:
                bin_names.append(">"+str(break_points[x]))
                bin_range.append((break_points[x],None))
    
        #add first and last label
        if break_points[0]>col_min:
            break_points.insert(0, -float("inf"))
        if break_points[-1]<col_max:
            break_points.append(float("inf"))
            
        #group values to bin
        column_bin = pd.cut(column, break_points, labels=bin_names)
        
        return column_bin, break_points, bin_names, bin_range
    
    def combine_segment_rules(self, input_df, column_name):
        rule_name_cols = [col for col in input_df.columns if column_name in col]
        input_df[column_name + '_combined'] = input_df[rule_name_cols].apply(lambda row: '|'.join(row.values.astype(str)), axis=1)
        input_df.drop(columns=rule_name_cols, inplace=True)
        
    def generate_segment_json(self):
        segment_config = json.loads('{}')
        
        #segment rules
        segment_rules_upload=self.segment_rules[['column','column_order','min_value','max_value','rule_name']]
        segment_rules_upload.reset_index(inplace=True)
        segment_rules_upload.rename(columns={'index':'segment_rule_id','column':'column_name'}, inplace=True)
        segment_rules_upload = segment_rules_upload.to_json(orient='records')
        
        #segment combined
        segments_upload = self.segment_rules_combined.drop(columns=['rule_eval_combined', 'rule_name_combined', 'column_combined']).copy()
        segments_upload.index.set_names('segment_id', inplace=True)
        segments_upload.reset_index(inplace=True)
        segments_upload_pivot = pd.melt(segments_upload, id_vars=['segment_id'], var_name='column_name', value_name='segment_rule_id') #'column_combined','rule_name_combined'
        segments_upload_pivot.dropna(inplace=True)
        segments_upload_pivot['segment_rule_id'] = segments_upload_pivot['segment_rule_id'].astype('int32')
        segments_upload = segments_upload_pivot[['segment_id','segment_rule_id']].to_json(orient='records')
        
        #return segment_rules_upload, segments_upload
        segment_config.update({'segment_input':self.segment_input})
        segment_config.update({'segment_rules':json.loads(segment_rules_upload)})
        segment_config.update({'segment_rules_combined':json.loads(segments_upload)})
        segment_config.update({'segment_curve_map':{}})
        
        return segment_config
    
class ModelConfig(object):
    """
    Class to import/create Model configuration settings including model execution
        commands, curve group to apply, amortization functions, cf modules to activate, interest rates, 
        payment matrixes, and more
    """
        
    def __init__(self, model_config_id, model_config_name, config_type=1, config_dict={}):
        
        self.model_config_id = model_config_id
        self.model_config_name = model_config_name
        self.config_type = config_type
        self.config_dict = config_dict
        
    def alter_config(self, metric_name, new_rule):
        self.config_dict[metric_name, new_rule]
        

class CurveStress(object):
    """
    Stress object to hold and create stress curve adjustments
    """
    
    def __init__(self, stress_dict):
        
        self.stress_name=''
        self.stress_config=stress_dict
        self.stress_df = {}
        
        for curve_type, stress in stress_dict.items():
        #    operator = "+" if stress[1]>0 else "-"
        #    stress_name = stress[0]+" "+operator+str(stress[1])+"%"
        #    self.stress_name = ';'.join([stress_name])
            
            self.add_stress(curve_type, stress)
    
    def add_stress(self, curve_type, stress):
        """
        Uses input instructions to create stress curves to apply in model
        
        Parameters
        =============================
        curve_type: str
            default, prepay, curtail
        stress: list
            list of tuples in format (period, stress %)
            will loop through each tuple as breakpoints to create stresses and scale values inbetween each point
        
            {'default': [(0, 0.10), (12, 0.10), (24, 0)],
             'prepay': [(0, -0.25), (12, -0.25), (24, 0)]
             }
        """
        
        stress_df_list = []
        
        #if list is length 1 apply stress to all months
        if len(stress)==1:
            stress_array = np.linspace(stress[0][1], stress[0][1], 300)
            stress_time= [x for x in range(0, 300)]
            stress_df_list.append(pd.DataFrame({'period':stress_time, 'rate':stress_array}, columns=['period', 'rate']))
            
        if len(stress)>1:
            #iterate 2 items at once
            for stress_begin, stress_end in zip(stress[:-1],stress[1:]):               
                stress_array = np.linspace(stress_begin[1], stress_end[1], (stress_end[0]-stress_begin[0]))
                stress_time = [x for x in range(stress_begin[0], stress_end[0])]
                stress_df_list.append(pd.DataFrame({'period':stress_time, 'rate':stress_array}, columns=['period', 'rate']))
            
        stress_df_combined = pd.concat(stress_df_list)
        stress_df_combined['period'] = stress_df_combined['period']+1
        stress_df_combined.set_index(['period'], inplace=True)
        
        self.stress_df[curve_type] = stress_df_combined
                    
    def return_stress_array(self, curve_types=[], max_length=360):

        output_arrays={}
        idx = range(max_length)
        
        for key, value in self.stress_df.items():
            if key in curve_types:
                reindex_df = value.reindex(idx, method='ffill', fill_value=0)
                reindex_array = reindex_df['rate'].to_numpy(dtype='float32')
                output_arrays[key] = reindex_array+1
            
        return output_arrays
        
class IndexProjection(object):
    """
    Class imports and stores interest rate forward projections.
    """
    
    def __init__(self, rate_projections, projection_date=None, cutoff_date=None):
        self.projection_date = projection_date
        self.cutoff_date  = cutoff_date
        
        self.index_list = {
            0:'Fixed',
            1:'1 Month LIBOR Rate',
            2:'2 Month LIBOR Rate',
            3:'3 Month LIBOR Rate',
            4:'6 Month LIBOR Rate',
            5:'1 Year LIBOR Rate',
            6:'PRIME',
            }
        
        self.raw_projections = rate_projections
        
        #format rate dataframe
        self.data_rate_curves = self.raw_projections.copy()
        self.data_rate_curves['curve_type'] = 'index'
        self.data_rate_curves['curve_sub_type'] = 'base'
        self.data_rate_curves['from_status'] = -1
        self.data_rate_curves['to_status'] = -1
        column_map = {'IndexName':'curve_id',
                      'ProjectionMonth':'period',
                      }
        self.data_rate_curves.rename(column_map, axis=1, inplace=True)
        self.data_rate_curves = self.data_rate_curves[['curve_type','curve_sub_type','curve_id', 'period','rate', 'from_status', 'to_status']].set_index(['curve_type','curve_sub_type','curve_id', 'period']).sort_index()
        

    
    
    
    
    
    
    