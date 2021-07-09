# -*- coding: utf-8 -*-
"""
Created on Thu Feb 13 15:18:46 2020

@author: jalbert
"""

import tkinter as tkr
from tkinter import ttk
import PySimpleGUI as sg
import openpyxl as xl
import re
import sys

import numpy as np
import pandas as pd

#from DataPrep.DataMover import ExcelEngine as xl_engine

#sg.change_look_and_feel('Dark Blue 3')
#sg.theme('Dark Blue 3')

class DataPrepMain(object):
    def __init__(self, data_prep):
        #store data_prep instance
        self.data_prep=data_prep
        #init objects
        self.import_data_tape = ImportDataTape(data_prep)
        self.import_rate_curves = ImportRateCurves(data_prep)
        self.map_segments = MapSegments(data_prep)

class MapSegments(object):
    """
    All GUI applications to collect user input
    """
    
    def __init__(self, data_prep):
        self.data_prep=data_prep
    
    def map_segments_main(self, curve_group):
        """
        GUI to map segments to curve options
        """
        segment_list = self.data_prep.segment_types.copy()
        #segment_map = self.data_prep.segment_curve_map.copy()
        segment_map = curve_group.segment_curve_map.copy()
        #unique_curves = self.data_prep.data_rate_curves.reset_index()[['curve_type','curve_id']].drop_duplicates()
        unique_curves = curve_group.curve_keys[['curve_name', 'curve_type']]
        
        all_tabs=[]
        
        for segment in segment_list:
            select_options = unique_curves[unique_curves['curve_type']==segment]['curve_name'].tolist()
            select_options.sort()
            selected_segment_map = segment_map[segment_map['segment_type']==segment].sort_values(by=['curve_name','segment_name'])
            
            row_cnt=0
            col_cnt=0
        
            tab_layout=[]        
            label_mstr=[]
            obj_mstr=[]
            row_mstr = []
            
            for ix, row in selected_segment_map.iterrows():
                col_cnt = 1 if col_cnt==3 else col_cnt+1
                #row_cnt+=1 if col_cnt==1 else 0
                
                box_desc = str(row['segment_name'])
                if str(row['curve_name']).strip()=='' or row['curve_name'] is None:
                    value = None
                else:
                    value=str(row['curve_name'])
                
                #label_mstr.append([sg.Text(box_desc+" :", size=(18,1), key=segment+'_label_'+box_desc, justification='right', pad=(1,3))])
                #obj_mstr.append([sg.Combo(select_options, default_value=value, key=segment+'_'+box_desc, size=(18,1), readonly=True, pad=(1,3))])
                row_mstr.append([sg.Text(box_desc+" :", size=(18,1), key=segment+'_label_'+box_desc, justification='right', pad=(1,3)), sg.Combo(select_options, default_value=value, key=segment+'_'+box_desc, size=(18,1), readonly=True, pad=(1,3))])
                #segment_row.extend([sg.Text(box_desc+" :", size=(20,1), key=segment+'_label_'+box_desc, justification='right'), sg.Combo(select_options, default_value=value, key=segment+'_'+box_desc, size=(20,1))])
                
                #if col_cnt==3:
                #    tab_layout.append(segment_row)
                #    segment_row=[]
            
            #on loop end append last row
            #if len(segment_row)>0:
            #    tab_layout.append(segment_row)
            
            #label_split = np.array_split(label_mstr,3)
            #obj_split = np.array_split(obj_mstr, 3)
            row_split = np.array_split(row_mstr, 3)
            
            for i in range(3):
                #tab_layout.extend([sg.Column(label_split[i], key=segment+'_lbl_column_'+str(i), element_justification='right')])
                #tab_layout.extend([sg.Column(obj_split[i], key=segment+'_obj_column_'+str(i), element_justification='left')])
                tab_layout.extend([sg.Column(row_split[i], key=segment+'_obj_column_'+str(i), element_justification='center')])
                if i<2:
                    tab_layout.extend([sg.VerticalSeparator()])
            
            
            all_tabs.append([sg.Tab(segment, key=segment, layout=([[sg.Column([tab_layout], key=segment+'_column', size=(1200,500), scrollable=True ,vertical_scroll_only=True, element_justification='left')]]))]) 
    
        layout=[
                [sg.Button("Save Selections")], #sg.Checkbox("Show Missing Only", key='show_checkbox', enable_events=True), 
                [sg.TabGroup(all_tabs, key='_tab_group')]
                ]
        
        window=sg.Window('Manual Curve Mapping', layout)
        while True:
            event, values = window.read()
            if event is None or event == 'Save Selections':
                break
            #if event=='show_checkbox':
            #    for key, value in window.AllKeysDict.items():
            #        if window[key].Type=='combo' or window[key].Type=='text':
            #            if window['show_checkbox'].get()==0:
            #                window[key].Update(visible=True)
            #                #window[key].set_size((20,0))
            #            elif window['show_checkbox'].get()==1 and values[key]!='':
            #                window[key].Update(visible=False)
                            #window[key].set_size((20,1))     
                                       
        #process dropdowns into dict of dicts (key for each segment)
        manual_map_dict={}
        for segment in segment_list:
            manual_map_dict[segment]={key[len(segment)+1:]:(value if str(value).strip()!='' else None) for key, value in values.items() if str(key).startswith(segment)}
            
        window.close()
        #self.data_prep.segment_map_manual = manual_map_dict
        curve_group.segment_map_manual= manual_map_dict
        #return manual_map_dict
    
class ImportDataTape(object):
    def __init__(self, data_prep):
        self.data_prep=data_prep

class ImportRateCurves(object):
    
    def __init__(self, data_prep):
        self.data_prep=data_prep
    
        
    def import_curves_main(self):
        
        #generate layout list for each segment type
        segment_types=['default','prepay', 'curtail', 'recovery','rollrate']
        loaded_files = {'':''}
        layout_main = []
        
        #segment toggles
        #for segment in segment_types:
            #layout_toggle.extend([sg.Frame(title=segment.capitalize(), layout=[[sg.Radio('None', group_id=segment+'_toggle', default=True, key=segment+'_toggle_none', enable_events=True)], [sg.Radio('SQL', group_id=segment+'_toggle', key=segment+'_toggle_sql', enable_events=True)], [sg.Radio('Excel', group_id=segment+'_toggle', key=segment+'_toggle_excel', enable_events=True)]])])    #[sg.Checkbox(text=str(segment).capitalize(), key=segment+'_toggle', enable_events=True)
            
        #layout_main.append([sg.Frame(layout=[layout_toggle], title='Select Curve Types to Load', key='segment_toggle_frame')])
        #segment_toggle = [sg.Column(layout=layout_toggle)]
        #layout_main.append([sg.Frame(layout=[layout_toggle], title='Select Curve Types to Load', key='segment_toggle_frame')])
        
        #excel loader options for each segment
        for segment in segment_types:
            #segment_toggle=[sg.Frame(title='Source', layout=[[sg.Radio('None', group_id=segment+'_toggle', default=True, key=segment+'_toggle_none', enable_events=True), sg.Radio('SQL', group_id=segment+'_toggle', key=segment+'_toggle_sql', enable_events=True), sg.Radio('Excel', group_id=segment+'_toggle', key=segment+'_toggle_excel', enable_events=True)]])]
            
            excel_frame_layout=[
                    [sg.Text('Select File'), sg.Combo(key=segment+'_file_nm_excel', values=list(loaded_files.keys()), default_value='', enable_events=True, size=(30,1), disabled=False, readonly=True), sg.FileBrowse(key=segment+'_file_selector_excel', disabled=False)],
                    [sg.Text('Select Worksheet'), sg.Combo((''), size=(20,1), key=segment+'_sheet_name_excel', disabled=False)], 
                    [sg.Text('Worksheet Range'), sg.InputText('', key=segment+'_ws_range_bgn_excel', disabled=False, size=(8,1)), sg.Text(':'), sg.InputText('', key=segment+'_ws_range_end_excel', disabled=False, size=(8,1))],
                    [sg.Checkbox('Unpivot Worksheet? (Column to Rows)', key=segment+'_unpivot_excel', enable_events=True, disabled=False)],
                    [sg.Text('Key Columns'), sg.InputText('', key=segment+'_pivot_keys_excel', tooltip='Comma delimited list, Identifier Columns to be used as the Curve Key', disabled=True)],
                    [sg.Button('Preview Data', key=segment+'_load_worksheet_excel', disabled=False)]
                    #[sg.Submit('View Worksheet'), sg.Cancel('Close')]
                    ]
            
            sql_frame_layout=[
                    [sg.Text('Select segment ID'), sg.Combo('')]
                    ]
            
            combined_frame_layout = [
                    #segment_toggle,
                    [sg.TabGroup([[sg.Tab('None', layout=[[sg.Text('No '+segment+' curves will be loaded', key=segment+'_none_text')]], key=segment+'_tab_none'),
                            sg.Tab('Excel', excel_frame_layout, key=segment+'_tab_excel'),
                            sg.Tab('SQL', sql_frame_layout, key=segment+'_tab_sql')]], key=segment+'_tabgroup')
                    ]]
            
            #combined_frame_layout = [[sg.Column(layout=excel_frame_layout), sg.VerticalSeparator(), sg.Column(layout=sql_frame_layout)]]
            layout_main.append(sg.Frame(layout=combined_frame_layout, title=str(segment).capitalize(), key=segment+'_frame', visible=True))
        
        #split into several rows
        row_length=3
        layout_main_split = [layout_main[x * row_length:(x+1) * row_length] for x in range((len(layout_main) + row_length - 1)//row_length)]
        

        #layout_main = [
        #        [sg.Frame(layout=[layout_toggle], title='Select Curve Types to Load', key='segment_toggle_frame')], 
                #[sg.Column(layout=layout_excel, scrollable=True, vertical_scroll_only=True)]#, title='Load from Excel', key='excel_load_frame')]
        #        [layout_excel]
        #        ]
                
        window = sg.Window('Import Rate Curves',layout_main_split)
        
        table_window_active = False
        
        while True:
            event, values = window.read()
            if event in (None, 'Close', 'Cancel'):
                break
            #if event[event.index('_')+1:].startswith('tab'):

            if event[event.index('_')+1:]=='file_nm_excel':
                segment=event[:event.index('_')]
                wb_path = values[event]
                
                if wb_path not in loaded_files.keys():
                    #load workbook
                    wb = self.data_prep.excel_engine.load_workbook(wb_path)
                    loaded_files[wb_path]=wb
                    #update sheet names for this segment
                    sheets = wb.sheetnames
                    window.FindElement(segment+'_sheet_name_excel').Update(values=sheets)
                    #update workbook path list for all segments
                    for key in window.AllKeysDict.keys():
                        if not isinstance(key,int) and key[key.index('_')+1:]=='file_nm_excel':
                            window.FindElement(key).Update(values=list(loaded_files.keys()))
                else:
                    #if wb is already loaded just populate sheet list for this segment
                    if wb_path=='':
                        window.FindElement(segment+'_sheet_name_excel').Update(values=[''])
                        window.FindElement(segment+'_sheet_name_excel').Update(value='')
                    else:
                        wb = loaded_files[wb_path]
                        sheets = wb.sheetnames
                        window.FindElement(segment+'_sheet_name_excel').Update(values=sheets)
                #make sure loaded workbook is selected for current segment
                window.FindElement(segment+'_file_nm_excel').Update(value=wb_path)
                
            if event[event.index('_')+1:].startswith('toggle'):
                segment=(event[:event.index('_')]).lower()
                
                #flip toggles
                if event.endswith('excel'):
                    pass
                    #window.FindElement(segment+'_)
                    #window.FindElement(segment+'_frame').update(visible=check_state)
                    #window.FindElement(segment+'_frame').unhide_row()
                    #for key in window.AllKeysDict.keys():
                    #    if key[:key.index('_')]==segment and key.endswith('excel') and key not in [segment+'_frame', segment+'_pivot_keys_excel']:
                    #        window.FindElement(key).Update(disabled=False)
                    #    elif key[:key.index('_')]==segment and 'toggle' not in key and key not in [segment+'_frame',segment+'_toggle']:
                    #        window.FindElement(key).Update(disabled=True)
                elif event.endswith('sql'):
                    pass
                    #for key in window.AllKeysDict.keys():
                    #    if key[:key.index('_')]==segment and key.endswith('sql') and key not in [segment+'_frame',segment+'_toggle', segment+'_pivot_keys_excel']:
                    #        window.FindElement(key).Update(disabled=False)
                    #    elif key[:key.index('_')]==segment and 'toggle' not in key and key not in [segment+'_frame']:
                    #        window.FindElement(key).Update(disabled=True)
                elif event.endswith('none'):
                    pass
                    #for key in window.AllKeysDict.keys():
                    #    if key[:key.index('_')]==segment and 'toggle' not in key and key not in [segment+'_frame', segment+'_pivot_keys_excel']:
                    #        window.FindElement(key).Update(disabled=True)
                            
            if event[event.index('_')+1:]=='unpivot_excel':
                window[segment+'_pivot_keys_excel'].Update(disabled=(not values[event]))
                
            if event[event.index('_')+1:]=='load_worksheet_excel':
                #try:
                    
                #collect values
                segment=event[:event.index('_')].lower()
                wb=values[(segment+'_file_nm_excel')]
                ws=values[(segment+'_sheet_name_excel')]
                
                bgn_rng = values[segment+'_ws_range_bgn_excel']
                end_rng = values[segment+'_ws_range_end_excel']
                
                unpivot_flag = values[(segment+'_unpivot_excel')]
                key_cols = values[segment+'_pivot_keys_excel']
                                
                #import data
                ws_data=self.data_prep.excel_engine.load_worksheet_range(loaded_files[wb][ws], bgn_rng, end_rng)
                ws_data.dropna(axis=1, how='all', inplace=True)
                #unpivot 
                if unpivot_flag:
                    ws_data = self.data_prep.excel_engine.unpivot_df(ws_data, pivot_col_name='period')
                #process key input
                key_cols = key_cols.strip().split(',')
                key_cols = [x.strip() for x in key_cols]
                if key_cols[0]!='': 
                    #combine key columns
                    curve_id = ws_data[key_cols].apply(lambda x: '|'.join(x.astype(str)), axis=1)
                    curve_name = '|'.join(key_cols)
                    ws_data.insert(0, 'curve_id',curve_id)
                    ws_data.insert(1, 'curve_name', curve_name)
                    #drop extra columns (anything used to create key)
                    keep_cols = ['curve_id','curve_name', 'period', 'rate'] #period_type, other_period_type,
                    #keep_cols.extend(key_cols)
                    remove_cols = [x for x in ws_data if x not in keep_cols]
                    ws_data.drop(columns=remove_cols, inplace=True)
                    #drop na rows
                    ws_data.dropna(axis=0, how='any', inplace=True)
                    #replace any non numeric values with zero
                    ws_data['rate']=(pd.to_numeric(ws_data['rate'], errors='coerce')).fillna(0)
                
                ws_data_sample = ws_data.iloc[:100, :25]
                ws_cols = ws_data_sample.columns
                
                #set column names if none exist
                col_names = ['column' + str(x) if ws_cols[x] is None else str(ws_cols[x]) for x in range(len(ws_cols))]
                
                table_layout=[[sg.Table(values=ws_data_sample.values.tolist(), headings=col_names, num_rows=25, display_row_numbers=True, vertical_scroll_only=False, auto_size_columns=False, max_col_width=25)]]
                table_window=sg.Window(segment.capitalize()+' Curve Data', table_layout)
                
                table_window_active=True
                
                #except:
                #    sg.PopupOK(sys.exc_info()) #'Cell range in this worksheet is invalid'
                #    table_window_active=False
                
            if table_window_active:
                ev_tbl, vals_tbl = table_window.Read(timeout=100)
                if ev_tbl is None or ev_tbl=='Exit':
                    table_window_active=False
                    table_window.close()
                
        window.close()
    





